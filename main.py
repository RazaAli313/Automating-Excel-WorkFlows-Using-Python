# openpyxl
# pip install openpyxl
from openpyxl.chart import BarChart,Reference
import openpyxl as xl


def process_spreadsheet(file):

    wb=xl.load_workbook(file)
    sheet=wb['Sheet1']
    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        updated_value=cell.value * 0.9
        new_cell=sheet.cell(row,4)
        new_cell.value=updated_value
    sheet.cell(1,4).value="Updated_Values"
    chart=BarChart()
    values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
    chart.add_data(values)
    sheet.add_chart(chart,"F2")
    newFile_name=file.replace(".xlsx","_updated.xlsx")
    wb.save(newFile_name)
    print(f"File Saved as {newFile_name}")
process_spreadsheet("sales.xlsx")
